#!/usr/bin/env python
# coding: utf-8

# In[1]:


#!/usr/bin/env python
# coding: utf-8

import pandas 
import matplotlib.pyplot as plt
import matplotlib
import seaborn
import os
from datetime import datetime
import warnings 
import numpy
from io import BytesIO
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import streamlit

matplotlib.rcParams.update({'savefig.bbox':'tight'})
pandas.options.mode.chained_assignment = None
pandas.set_option('display.max_rows', None)

warnings.filterwarnings("ignore", category=RuntimeWarning)

now = datetime.today().strftime('%Y%m%d_%H%M%S')

# ALL FUNCTIONS

# In[2]:


def prepare_excel_download(file_path: str, label: str = "Download Excel File"):
    """
    Prepares an Excel file for download in a Streamlit app.
    """
    with open(file_path, "rb") as f:
        file_bytes = f.read()

    # portable way to extract filename
    file_name = os.path.basename(file_path)

    streamlit.download_button(
        label=label,
        data=file_bytes,
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# In[9]:


streamlit.title('ddPCR HDR Analysis Automation')
streamlit.header('Directions')
streamlit.text('Upload a QX200 Droplet Reader Excel file and a sample key in the drop boxes below to produce a downloadable Excel file. Optionally, set the parameters below prior to uploading your files to customize the analysis.')
streamlit.title('Inputs')
streamlit.header('QX200 File')
fileName = streamlit.file_uploader('Upload the QX200 output excel file here', 'xlsx')
streamlit.header('Sample Key')
streamlit.text('Sample key must be filled out with sample names separated by underscores, where the last field in the sample name indicates the replicate. Replicates need to be named identically, with the exception of this last field.  For example, replicate names of "LKP23059_MOI40k_Rep1" and "LKP23059_MOI40k_Rep2", "LKP23059_1" and "LKP23059_2", as well as "LKP23059_4.58RNP_40kMOI_A" and "LKP23059_4.58RNP_40kMOI_B" are all acceptable. The number of fields does not matter so long as the part that indicates the replicate comes last. You can leave the "Sample Entity Link" and "Analytical Control Link" fields blank if they are not needed.')
streamlit.text('Download the Sample Key Template below')
prepare_excel_download('SampleKeyTemplate.xlsx', 'Download Key Template')
keyName = streamlit.file_uploader('Upload your filled in sample key here', 'xlsx')

streamlit.header('Optional Parameters')
streamlit.text('These are set to default values, but change them to customize your analysis!')

# user-provided output file name
raw_outFile = streamlit.text_input('Enter the desired name of your output Excel file',
                                   value='25BCPXXX_AnalyzedResults.xlsx')

# ----------------- SAFE PATH HANDLING -----------------
out_dir = Path("Results")
out_dir.mkdir(parents=True, exist_ok=True)

raw = (raw_outFile or "").strip()
basename = Path(raw).name  # just the filename

if not basename:
    basename = f"AnalyzedResults_{now}.xlsx"

# strip invalid characters for Windows
invalid_chars = '<>:"/\\|?*'
basename = ''.join("_" if c in invalid_chars else c for c in basename)

if not basename.lower().endswith(".xlsx"):
    basename += ".xlsx"

outPath = str(out_dir / basename)

streamlit.text(f"Output Excel will be written to: {outPath}")
# ------------------------------------------------------

famName = streamlit.text_input('What is the name of your FAM target in your input Excel sheet?', value='CCR5')
hexName = streamlit.text_input('What is the name of your HEX target in your input Excel sheet?', value='CCRL2')
cvThresh = streamlit.slider('Replicate CVs above this value will be highlighted in red',
                            min_value=0, max_value=50, value=5)

streamlit.title('Results')




# In[10]:


def color_code_column_heatmap(df, column_name, output_file=outPath):
    """
    Write a DataFrame to Excel and color-code cells in a given column using a softer
    red→green pastel heatmap.
    """

    # Save DataFrame to Excel first
    df.to_excel(output_file, index=False)

    # Load workbook with openpyxl
    wb = load_workbook(output_file)
    ws = wb.active

    # Get column index for the target column
    col_idx = None
    for idx, col in enumerate(ws[1], start=1):
        if col.value == column_name:
            col_idx = idx
            break

    if col_idx is None:
        raise ValueError(f"Column '{column_name}' not found in DataFrame")

    # Extract column values from DataFrame to compute min and max
    values = df[column_name].dropna().astype(float)
    min_val, max_val = values.min(), values.max()

    def blend_with_white(r, g, b, blend=0.5):
        """Blend an RGB color with white (255,255,255) to soften it"""
        r = int(r + (255 - r) * blend)
        g = int(g + (255 - g) * blend)
        b = int(b + (255 - b) * blend)
        return r, g, b

    def value_to_color(val):
        """Map value to a softer red→green pastel hex color"""
        ratio = (val - min_val) / (max_val - min_val) if max_val > min_val else 0.5

        # Strong base color
        r = int(255 * (1 - ratio))
        g = int(255 * ratio)
        b = 0

        # Blend with white to make pastel
        r, g, b = blend_with_white(r, g, b, blend=0.6)  # 0.6 = more white → softer

        return f"{r:02X}{g:02X}{b:02X}"

    # Loop through column values (skip header row)
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        try:
            val = float(cell.value)
        except (TypeError, ValueError):
            continue  # skip non-numeric

        hex_color = value_to_color(val)
        cell.fill = PatternFill(start_color=hex_color,
                                end_color=hex_color,
                                fill_type="solid")

    # Save workbook
    wb.save(output_file)



def insert_png_into_excel(excel_file, png_path, cell, output_file=None, sheet_name=None):
    """
    Insert a PNG into an Excel sheet using openpyxl.

    Args:
        excel_file (str): Path to the Excel file to modify (or None to create a new one).
        png_path (str): Path to the PNG image file.
        cell (str): Cell reference (e.g. "B2") where to insert the image.
        output_file (str): File path to save. If None, overwrites excel_file.
        sheet_name (str): Sheet to insert into. If None, uses the active sheet.
    """

    # Load or create workbook
    if excel_file and excel_file.endswith(".xlsx"):
        wb = load_workbook(excel_file)
    else:
        wb = Workbook()

    # Select sheet
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    # Load image
    img = Image(png_path)

    # Add to worksheet at specified cell
    ws.add_image(img, cell)

    # Save
    if output_file is None:
        output_file = excel_file or "output_with_image.xlsx"
    wb.save(output_file)

    


def color_code_column_red(excel_file, column_name, threshold, output_file=None, sheet_name=None):
    """
    Open an existing Excel file and color-code cells in a given column if values exceed a threshold.

    Args:
        excel_file (str): Path to the existing Excel file
        column_name (str): Name of the column to color code
        threshold (float or int): Threshold value for coloring
        output_file (str): Path to save Excel file (if None, overwrites original)
        sheet_name (str): Sheet to edit (if None, uses active sheet)
    """

    # Load workbook
    wb = load_workbook(excel_file)

    # Select sheet
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    # Get column index for the target column
    col_idx = None
    for idx, col in enumerate(ws[1], start=1):
        if col.value == column_name:
            col_idx = idx
            break

    if col_idx is None:
        raise ValueError(f"Column '{column_name}' not found in sheet '{ws.title}'")

    # Loop through column values (skip header row)
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        try:
            val = float(cell.value)
        except (TypeError, ValueError):
            continue  # skip non-numeric

        if val > threshold:
            # Create a color shade based on how much it exceeds threshold
            intensity = 255
            g_b = 100
            argb = f"FF{intensity:02X}{g_b:02X}{g_b:02X}"  # reddish gradient
     
            cell.fill = PatternFill(start_color=argb,
                                    end_color=argb,
                                    fill_type="solid")

    # Save workbook
    if output_file is None:
        output_file = excel_file  # overwrite if not provided
        wb.save(output_file)
        

def seaborn_download_button(fig, filename="plot.png", label="Download plot"):
    """
    Create a Streamlit download button for a Seaborn/Matplotlib figure.
    
    Parameters:
        fig (matplotlib.figure.Figure): The Seaborn/Matplotlib figure object.
        filename (str): The default filename for the downloaded file.
        label (str): The text displayed on the download button.
    """
    buf = BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight")
    buf.seek(0)
    streamlit.download_button(
        label=label,
        data=buf,
        file_name=filename,
        mime="image/png"
    )

def pass_fail_crit(df, group, famName, hexName, copyThresh=20, dropThresh=10000, cvThresh=5):
    
    df = df.reset_index().set_index(['Sample Group', 'Target']).sort_index()
    famFlag = 0
    hexFlag = 0
    copyFlag = 0
    dropFlag = 0
    cvFlag = 0

    
    if df.loc[(group, famName), 'Conc(copies/µL)'].mean() >= copyThresh:
        famFlag =1
    if df.loc[(group, hexName), 'Conc(copies/µL)'].mean() >= copyThresh:
        hexFlag =1
    if df.reset_index().set_index(['Sample Group']).sort_index().loc[group, 'Accepted Droplets'].mean() >= dropThresh:
        dropFlag = 1
    if df.reset_index().set_index(['Sample Group']).sort_index().loc[group, 'Replicate CV'].mean() <= cvThresh: 
        cvFlag = 1
        
    if famFlag==1 and hexFlag==1 and dropFlag == 1 and cvFlag == 1:
        return 'Pass'
    else:
        return 'Fail'
    
    

#####################################################################################################


# In[11]:


# In[3]:


if fileName and keyName:
    inFile = pandas.read_excel(fileName)
    key = pandas.read_excel(keyName)
    impCols = inFile[['Well', 'Target', 'Conc(copies/µL)', 'Accepted Droplets', 'Positives', 'Negatives']]
    impCols['Target'] = impCols['Target'].astype(str)
    impCols['Name'] = ''
    impCols['Sample Group'] = ''
    impCols['Sample Entity Link'] = ''
    impCols['Analytical Control Link'] = '' 
    key.set_index('Well', inplace = True)
    impCols.set_index('Well', inplace = True)
    for well in impCols.index:
        impCols.loc[well, 'Name'] = key.loc[well, 'Name']
        try:
            impCols['Sample Entity Link'] = impCols['Sample Entity Link'].fillna('')
            impCols.loc[well, 'Sample Entity Link'] = key.loc[well, 'Sample Entity Link']
        except KeyError:
            impCols.loc[well, 'Sample Entity Link'] = ''
        try:
            impCols['Analytical Control Link'] = impCols['Analytical Control Link'].fillna('')
            impCols.loc[well, 'Analytical Control Link'] = key.loc[well, 'Analytical Control Link']
        except KeyError:
            impCols.loc[well, 'Sample Entity Link'] = ''
        groupList = key.loc[well, 'Name'].split('_')[:-1]
        impCols.loc[well, 'Sample Group'] = '_'.join(groupList)     


    # In[4]:


    impCols.reset_index(inplace = True)
    impCols.set_index('Name', inplace = True)
    impCols.sort_index(inplace = True)

    samples = impCols.index
    impCols.reset_index(inplace = True)
    sortLevelList = ['Name', 'Target']
    impCols.set_index(sortLevelList, inplace = True)
    impCols['HDR(%)'] = float('nan')


    # In[5]:


    for sample in samples:
        ccr5 = [sample]+[famName]
       
        ccrl2 = [sample]+[hexName]
        try:
            hdr = 100*impCols.loc[tuple(ccr5), 'Conc(copies/µL)'] / impCols.loc[tuple(ccrl2), 'Conc(copies/µL)']
            #print(hdr)
        except TypeError: 
            hdr = 0
        impCols.loc[sample, 'HDR(%)'] = hdr


    # In[6]:


    impCols.reset_index(inplace = True)
    impCols.set_index('Sample Group', inplace = True)
    repGroups = impCols.index


    # In[7]:


    impCols['Avg. HDR(%)'] = float('nan')
    impCols['Replicate CV'] = float('nan')
    graphingFrame = pandas.DataFrame(columns = ['Group', 'HDR'])
    benchlingFrame = pandas.DataFrame(columns = ['Sample ID (Text)', 'Sample Entity Link', 'Analytical Control Link', '% Targeted Integration', 'Copies/uL (FAM)', 'Copies/uL (HEX)', 'Droplet Number', '%CV (integration)', 'Pass/Fail', 'Primer Target (FAM)', 'Primer Target (HEX)'])
    iters = 0

    impCols['Conc(copies/µL)'] = impCols['Conc(copies/µL)'].replace('No Call', 0)
    for group in repGroups: 

        avghdr = impCols.loc[group, 'HDR(%)'].mean()
        cv = 100*impCols.loc[group, 'HDR(%)'].std()/impCols.loc[group, 'HDR(%)'].mean()
        if numpy.isnan(cv):
            cv = 0

        impCols.loc[group, 'Avg. HDR(%)'] = avghdr
        impCols.loc[group, 'Replicate CV'] = cv
        
        graphingFrame.loc[iters, 'Group'] = group
        graphingFrame.loc[iters, 'HDR'] = avghdr
        
        benchlingFrame.loc[iters, 'Sample ID (Text)'] = group
        benchlingFrame.loc[iters, 'Sample Entity Link'] = impCols.loc[group, 'Sample Entity Link'].drop_duplicates().values
        benchlingFrame.loc[iters, 'Analytical Control Link'] = impCols.loc[group, 'Analytical Control Link'].drop_duplicates().values
        benchlingFrame.loc[iters, '% Targeted Integration'] = avghdr 
        benchlingFrame.loc[iters, 'Copies/uL (FAM)'] = impCols.reset_index().set_index(['Sample Group', 'Target']).sort_index().loc[(group, famName), 'Conc(copies/µL)'].mean()
        benchlingFrame.loc[iters, 'Copies/uL (HEX)'] = impCols.reset_index().set_index(['Sample Group', 'Target']).sort_index().loc[(group, hexName), 'Conc(copies/µL)'].mean()
        benchlingFrame.loc[iters, 'Droplet Number'] = impCols.loc[group, 'Accepted Droplets'].mean()
        benchlingFrame.loc[iters, '%CV (integration)'] = cv
        benchlingFrame.loc[iters, 'Pass/Fail'] = pass_fail_crit(impCols, group, famName, hexName, 20, 10000, cvThresh)
        benchlingFrame.loc[iters, 'Primer Target (FAM)'] = famName
        benchlingFrame.loc[iters, 'Primer Target (HEX)'] = hexName
        
        iters+=1

    #impCols.to_excel(outPath)
    graphingFrame = graphingFrame.drop_duplicates('Group').sort_values(by = 'Group').reset_index(drop = True)
   


    # In[8]:


    barfig = plt.figure()
    seaborn.barplot(data = graphingFrame, x = 'Group', y = 'HDR', palette = 'flare', dodge = False)
    plt.title('HDR Values')
    plt.xticks(rotation = 90, fontsize=8)
    plt.legend(loc = 'upper right', bbox_to_anchor=[1.20,1.00])
    plt.ylabel('HDR Alleles (%)')
    plt.savefig(os.path.join('Results', now +'_HDR_BarGraph.png'))


    # In[4]:


    dropfig = plt.figure()
    seaborn.scatterplot(data = impCols, x = 'Positives', y = 'Negatives', hue = 'Target', palette = 'cool')
    plt.xlabel('Positive Droplets')
    plt.ylabel('Negative Droplets')
    plt.title('Droplet Stats')
    plt.xlim([0, 25000])
    plt.ylim([0, 25000])
    plt.savefig(os.path.join('Results',  now+'_DropletStats.png'))


    # In[11]


    color_code_column_heatmap(impCols, 'Avg. HDR(%)')
    
    insert_png_into_excel(outPath, os.path.join('Results', now+'_HDR_BarGraph.png'), 'Q2')
    insert_png_into_excel(outPath, os.path.join('Results', now+'_DropletStats.png'), 'Q30')
    color_code_column_red(outPath, 'Replicate CV', cvThresh)

    benchlingFrame = benchlingFrame.drop_duplicates('Sample ID (Text)').reset_index(drop = True)
    with pandas.ExcelWriter(outPath, engine="openpyxl", mode="a") as writer:
        benchlingFrame.to_excel(writer, sheet_name="Benchling Queryable Output", index=False)

    wb = load_workbook(outPath)
    sheet = wb["Benchling Queryable Output"]

    # Move new sheet to the first position
    wb._sheets.insert(0, wb._sheets.pop(wb._sheets.index(sheet)))

    wb.save(outPath)


    streamlit.header('Analyzed Outputs')

    prepare_excel_download(outPath)
    
    

if keyName and fileName:
    streamlit.text('Bar plot of average HDR values, plotted on a per sample basis.') 
    streamlit.pyplot(barfig)
    seaborn_download_button(barfig, filename = "HDRBarGraph.png", label = 'Donwload Bar Graph')
    streamlit.text('Scatter plot of droplet contents, broken out by target.')
    streamlit.pyplot(dropfig)
    seaborn_download_button(dropfig, filename = 'DropletsScatterPlot.png', label = 'Download Scatter Plot')


# In[ ]:




